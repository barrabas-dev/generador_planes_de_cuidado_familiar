import os
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image


def extraer_datos_excel(
    ruta_excel,
    nombre_hoja="Hoja1",
    carpeta_fam="assets/imagenes/familiogramas",
    carpeta_eco="assets/imagenes/ecomapas",
    carpeta_ana="assets/analisis"
):
    """
    Extrae imágenes de una hoja Excel (columnas B y C) y texto (columna D),
    guardándolos en carpetas específicas según su ubicación.
    """
    # Crear carpetas de destino
    for carpeta in (carpeta_fam, carpeta_eco, carpeta_ana):
        os.makedirs(carpeta, exist_ok=True)

    # Cargar el archivo Excel
    wb = load_workbook(ruta_excel, data_only=True)
    hoja = wb[nombre_hoja]

    # Asociar imágenes a sus coordenadas de celda
    imagenes_por_celda = {}
    for img in hoja._images:  # openpyxl no ofrece acceso público directo
        coord = (img.anchor._from.col + 1, img.anchor._from.row + 1)  # (col, fila)
        imagenes_por_celda[coord] = img

    # Procesar cada fila (desde la fila 2, suponiendo encabezado en la 1)
    for fila in hoja.iter_rows(min_row=2):
        id_hogar = str(fila[0].value).strip()

        fila_idx = fila[0].row
        coord_fam = (2, fila_idx)  # columna B
        coord_eco = (3, fila_idx)  # columna C
        celda_ana = hoja.cell(row=fila_idx, column=4)  # columna D

        # Guardar imágenes si existen
        if coord_fam in imagenes_por_celda:
            guardar_imagen(imagenes_por_celda[coord_fam], os.path.join(carpeta_fam, f"{id_hogar}.jpg"))

        if coord_eco in imagenes_por_celda:
            guardar_imagen(imagenes_por_celda[coord_eco], os.path.join(carpeta_eco, f"{id_hogar}.jpg"))

        # Guardar texto de análisis en .txt
        if celda_ana.value:
            guardar_texto(celda_ana.value, os.path.join(carpeta_ana, f"{id_hogar}.txt"))

    print("✅ Proceso completado con éxito.")


def guardar_imagen(img_obj, ruta_destino):
    """
    Guarda una imagen de Excel en formato JPG.
    """
    if hasattr(img_obj, "_data"):
        data = img_obj._data()
    else:
        with open(img_obj.ref, "rb") as f:
            data = f.read()

    image = Image.open(io.BytesIO(data))
    if image.mode != "RGB":
        image = image.convert("RGB")
    image.save(ruta_destino, "JPEG", quality=95)


def guardar_texto(texto, ruta_destino):
    """
    Guarda texto plano en un archivo .txt.
    """
    with open(ruta_destino, "w", encoding="utf-8") as f:
        f.write(str(texto).strip())


if __name__ == "__main__":
    extraer_datos_excel("assets/imagenes.xlsx", nombre_hoja="Hoja 1")
