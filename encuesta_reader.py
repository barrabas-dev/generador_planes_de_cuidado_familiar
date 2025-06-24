# encuesta_reader.py

from typing import List, Dict
from dataclasses import dataclass, field
import openpyxl


# Diccionario base de columnas
encuesta_dict = {
    "HOGAR": None,
    "NOMBRE": None,
    "TIPO DE ID": None,
    "NUMERO DE IDENTIFICACIÓN": None,
    "GENERO": None,
    "EDAD": None,
    "EPS": None,
    "REGIMEN": None,
    "FECHA DE NACIMIENTO": None,
    "ESTADO CIVIL": None,
    "ESCOLARIDAD": None,
    "OCUPACION": None,
    "SELECCIONADO": None,
    "CONVIVE DENTRO DE LA CASA": None,
    "RUTA A LA QUE PERTENECE": None,
    "TELEFONO": None,
    "Direccion": None,
    "CITOLOGÍA- ADN VPH": None,
    "MAMOGRAFIA - ECM": None,
    "TAMIZAJE CA DE COLON": None,
    "TAMIZAJE CA DE PROSTATA": None,
    "DESPARASITACION ANTIHELMITICA": None,
    "PLANIFICACION FAMILIAR": None,
    "TAMIZAJE ANEMIA Y HEMOGLOBINA": None,
    "TAMIZAJE RIESGO CARDIOVASCULAR": None,
    "VACUNACIÓN": None,
    "VALORACIÓN ODONTOLOGÍA": None,
    "CONSULTA DE CONTROL (RIAS)": None,
    "TOMA DE LABORATORIOS SEGUN RIA": None
}


# Campos de tipo booleano interpretado
BOOLEAN_COLUMNS = {

    "SELECCIONADO",
    "CITOLOGÍA- ADN VPH",
    "MAMOGRAFIA - ECM",
    "TAMIZAJE CA DE COLON",
    "TAMIZAJE CA DE PROSTATA",
    "DESPARASITACION ANTIHELMITICA",
    "PLANIFICACION FAMILIAR",
    "TAMIZAJE ANEMIA Y HEMOGLOBINA",
    "TAMIZAJE RIESGO CARDIOVASCULAR",
    "VACUNACIÓN",
    "VALORACIÓN ODONTOLOGÍA",
    "CONSULTA DE CONTROL (RIAS)",
    "TOMA DE LABORATORIOS SEGUN RIA"
}


@dataclass
class Paciente:
    datos: Dict[str, str | bool]


@dataclass
class NucleoFamiliar:
    hogar_id: str
    integrantes: List[Paciente] = field(default_factory=list)


def interpretar_valor(valor: str, columna: str) -> str | bool:
    """
    Interpreta los valores booleanos según las reglas del usuario.
    Si el valor está vacío -> False.
    Si contiene la palabra "NO" -> True.
    Si no es booleano -> retorna el mismo string.
    """
    if columna in BOOLEAN_COLUMNS:
        if not valor or valor.strip() == "":
            return False
        elif "NO" or "SI" in valor.upper():
            return True
        else:
            return False
    return valor.strip()


def leer_encuesta(path_excel: str, hoja: str = None) -> Dict[str, NucleoFamiliar]:
    """
    Lee la hoja de Excel y organiza los datos por núcleos familiares.
    """
    wb = openpyxl.load_workbook(path_excel, data_only=True)
    ws = wb[hoja] if hoja else wb.active

    nucleos: Dict[str, NucleoFamiliar] = {}

    # Fila 3 contiene los encabezados
    encabezados = [cell.value for cell in ws[3]]

    # Recorremos desde la fila 4 en adelante
    for row in ws.iter_rows(min_row=4, values_only=True):
        paciente_data = {}
        hogar_id = None

        # Asociamos los datos de la fila con los encabezados
        for idx, valor in enumerate(row):
            columna = encabezados[idx]
            if columna in encuesta_dict:
                interpretado = interpretar_valor(str(valor) if valor is not None else "", columna)
                paciente_data[columna] = interpretado
                if columna == "HOGAR":
                    hogar_id = str(interpretado)

        # Saltamos si no hay hogar_id
        if not hogar_id:
            continue

        paciente = Paciente(datos=paciente_data)

        # Agrupamos por núcleo familiar
        if hogar_id not in nucleos:
            nucleos[hogar_id] = NucleoFamiliar(hogar_id=hogar_id)

        nucleos[hogar_id].integrantes.append(paciente)

    return nucleos
