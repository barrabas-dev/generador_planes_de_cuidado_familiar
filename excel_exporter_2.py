import os
import shutil
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment
from math import ceil
from openpyxl.utils import get_column_letter, column_index_from_string
from factores_encontrados import obtener_acuerdo_completo_apgar, obtener_fortalezas_apgar, generar_textos_factores

def crear_planes_cuidado_familiares(dato_nucleos: dict, planes_nucleos: dict, plantilla_path: str, salida_dir: str):
    """
    Crea un archivo Excel por cada núcleo familiar incluido en 'planes_nucleos'.
    Cada archivo contiene una única hoja, basada en una plantilla, con la información de:
    - Factores del plan familiar (desde 'planes_nucleos')
    - Miembros del núcleo (desde 'dato_nucleos')

    :param dato_nucleos: Diccionario con todos los núcleos y sus integrantes.
    :param planes_nucleos: Diccionario con los núcleos seleccionados y sus factores.
    :param plantilla_path: Ruta al archivo Excel plantilla.
    :param salida_dir: Carpeta donde se guardarán los archivos generados.
    """

    if not os.path.isfile(plantilla_path):
        raise FileNotFoundError(f"Plantilla no encontrada en la ruta: {plantilla_path}")

    os.makedirs(salida_dir, exist_ok=True)

    for nucleo_id, plan_info in planes_nucleos.items():
        if nucleo_id not in dato_nucleos:
            print(f"⚠️  Núcleo {nucleo_id} no encontrado en datos de integrantes. Se omite.")
            continue

        integrantes = dato_nucleos[nucleo_id]["integrantes"]

        # Crear copia del archivo base
        salida_path = os.path.join(salida_dir, f"{nucleo_id}.xlsx")
        shutil.copy(plantilla_path, salida_path)

        wb = load_workbook(salida_path)
        hoja = wb.worksheets[0]
        hoja.title = nucleo_id


        # Obtener info del plan
        fecha_tupla = plan_info.get("fecha_visita")
        resultado_apgar = plan_info.get("resultado_apgar")
        acuerdo_apgar = (obtener_acuerdo_completo_apgar(resultado_apgar) or "BOlO").strip()


        # Llenar información en la hoja
        anexar_texto(hoja, "X8", nucleo_id, ajustar_altura=False)
        anexar_texto(hoja, "B8", "Toledo", ajustar_altura=False)
        anexar_texto(hoja, "H8", "Santa Lucia", ajustar_altura=False) #para modificar barrio o vereda
        anexar_fecha_en_celdas(hoja, "C10", fecha_tupla)
        anexar_texto(hoja, "B11", "EBAS GRUPO 3 URBANO", ajustar_altura=False)
        anexar_texto(hoja, "C16", resultado_apgar, ajustar_altura=False)
        anexar_texto(hoja,"E17", acuerdo_apgar, ancho_estimado=70)

        # Escribir fortalezas en las celdas especificadas
        celdas_destino_fortalezas = ("B30", "B32", "M30", "M32")
        fortalezas = obtener_fortalezas_apgar(resultado_apgar)
        escribir_fortalezas_en_celdas(hoja, celdas=celdas_destino_fortalezas, fortalezas=fortalezas)

        # Escribir respuestas por hallazgos cuidado en salud familiar
        from db_restuestas import factores_dict
        celdas_destino_cuidado_hallazgos = ("C37", "I37", "U37", "V37")
        cuidado_hallazgos = generar_textos_factores(plan_info, factores_dict)
        escribir_duidado_por_hallazgos_en_celdas(hoja, celdas=celdas_destino_cuidado_hallazgos, respuesta_hallazgos=cuidado_hallazgos)
       
        #Llenar cuidado por curso de vida
        from factores_encontrados import generar_tuplas_integrantes
        from db_restuestas import datos_por_ciclo_vida

        tuplas = generar_tuplas_integrantes(integrantes, datos_por_ciclo_vida)
        rellenar_celdas_integrantes(hoja, tuplas)


        #for t in tuplas: #ciclo para ver en consula si funciona la funcion
            #print(t)


        # Llenar nombre y datos de la primera persona del núcleo en B13
        if integrantes:
            primer_nombre = integrantes[0].get("NOMBRE", "")
            telefono = integrantes[0].get("TELEFONO", "")
            direccion = integrantes[0].get("Direccion","")
            anexar_texto(hoja, "B13", primer_nombre, ajustar_altura=False)
            anexar_texto(hoja, "J13", primer_nombre, ajustar_altura=False)
            anexar_texto(hoja, "U13", telefono, ajustar_altura=False)
            anexar_texto(hoja, "N8", direccion, ajustar_altura=False)

        else:
            print(f"⚠️ Núcleo {nucleo_id} no tiene integrantes. B13 quedará sin nombre.")



        #anexar_texto(hoja["B4"], fecha_visita)
        #anexar_texto(hoja["B5"], resultado_apgar)
        '''
        # Llenar información en la hoja
        hoja["X8"] = nucleo_id  # Ejemplo de celda para el ID del núcleo
        hoja["B4"] = fecha_visita
        hoja["B5"] = resultado_apgar

        '''

        wb.save(salida_path)
        print(f"✅ Plan generado para núcleo {nucleo_id}: {salida_path}")
        #print("Valor:", resultado_apgar)
        #print(integrantes.get("EDAD",""))

        for i, persona in enumerate(integrantes, start=1):
            edad = persona.get("EDAD", "No especificada")
            print(f"Integrante {i} - Edad: {edad}")


#--------------------V0 de la funcion-----------------------------
''' 
def anexar_texto(hoja, ref, nuevo_valor, ancho_estimado):
    """
    Anexa texto a una celda respetando su valor actual. 
    Si la celda está en una región combinada, se edita solo la celda superior izquierda.
    Ajusta automáticamente la altura de la fila según el contenido.

    :param hoja: Objeto hoja de openpyxl.
    :param ref: Referencia como "X8".
    :param nuevo_valor: Texto a agregar.
    :param ancho_estimado: Cantidad estimada de caracteres por línea visible en la celda.
    """
    celda = hoja[ref]

    # Si la celda es parte de un rango combinado, ubicamos la principal
    if isinstance(celda, MergedCell):
        for rango in hoja.merged_cells.ranges:
            if ref in rango:
                ref = rango.coord.split(":")[0]  # Ej: "X8"
                celda = hoja[ref]
                break

    texto_existente = celda.value or ""
    nuevo_texto = f"{texto_existente} {nuevo_valor}".strip()
    celda.value = nuevo_texto

    # Ajuste automático de altura de la fila
    fila = celda.row
    num_lineas = len(nuevo_texto) // ancho_estimado + 1
    alto_por_linea = 15  # puedes ajustar este valor si ves que se corta o queda muy alto

    hoja.row_dimensions[fila].height = num_lineas * alto_por_linea
'''
#------------------V1 de la funcion anexar_texto-----------------------------------

from math import ceil
from openpyxl.cell.cell import MergedCell

def anexar_texto(hoja, ref, nuevo_valor, ajustar_altura=True, ancho_estimado=60):
    """
    Anexa texto a una celda respetando su valor actual. 
    Si la celda está en una región combinada, se edita solo la celda superior izquierda.
    Además, ajusta opcionalmente la altura de la fila según el contenido.

    :param hoja: Objeto de hoja de openpyxl.
    :param ref: Referencia a la celda como "X8".
    :param nuevo_valor: Texto a agregar.
    :param ajustar_altura: Booleano, si True ajusta la altura de la celda automáticamente.
    :param ancho_estimado: Ancho estimado en caracteres para calcular el alto si se ajusta.
    """

    '''
    Uso de la funcion

    # Solo agregar texto, sin modificar alto
    anexar_texto(hoja, "C10", "Texto simple", ajustar_altura=False)

    # Agregar texto y ajustar altura (ancho por defecto de 60 caracteres)
    anexar_texto(hoja, "D12", "Texto largo para ajustar la altura automáticamente")

    # Agregar texto y ajustar altura con un ancho más generoso (menos alto)
    anexar_texto(hoja, "E15", "Otro texto extenso", ancho_estimado=80)

    '''
    celda = hoja[ref]

    # Si la celda es parte de un rango combinado, buscamos la celda editable principal
    if isinstance(celda, MergedCell):
        for rango in hoja.merged_cells.ranges:
            if ref in rango:
                ref = rango.coord.split(":")[0]  # ej. "X8"
                celda = hoja[ref]
                break

    # Anexar el texto al contenido actual
    texto_existente = celda.value or ""
    celda.value = f"{texto_existente} {nuevo_valor}".strip()


    # Activar ajuste de texto si no está activo (ojito)
    alineacion_actual = celda.alignment or Alignment()
    if not alineacion_actual.wrapText:
        celda.alignment = Alignment(
            horizontal=alineacion_actual.horizontal,
            vertical=alineacion_actual.vertical or "top",
            wrapText=True
        )


    # Ajustar altura de la fila según el contenido (opcional)
    if ajustar_altura:
        texto = celda.value or ""
        num_lineas = ceil(len(texto) / ancho_estimado)
        altura_estim = max(15, num_lineas * 15)  # altura mínima de 15
        hoja.row_dimensions[celda.row].height = altura_estim




#funcion para anexar fecha separada
def anexar_fecha_en_celdas(hoja, celda_inicio: str, fecha_tupla: tuple):
    """
    Recibe una hoja de Excel, una celda de inicio (como 'C10') y una tupla con fecha (día, mes, año).
    Escribe los valores en celdas consecutivas hacia la derecha.
    """
    try:
        dia, mes, anio = fecha_tupla
    except (ValueError, TypeError):
        print(f"[ERROR] La fecha no es una tupla válida: {fecha_tupla}")
        dia, mes, anio = "", "", ""

    # Extraer fila y columna de la celda de inicio
    fila = int(''.join(filter(str.isdigit, celda_inicio)))
    col_letra = ''.join(filter(str.isalpha, celda_inicio))
    col_index = column_index_from_string(col_letra)

    # Escribir valores en celdas consecutivas
    hoja[f"{get_column_letter(col_index)}{fila}"] = dia
    hoja[f"{get_column_letter(col_index + 1)}{fila}"] = mes
    hoja[f"{get_column_letter(col_index + 2)}{fila}"] = anio

#---------------Funcion para añadir fortalezas a plantilla---------------------

from openpyxl.worksheet.worksheet import Worksheet
from typing import Tuple

def escribir_fortalezas_en_celdas(
    hoja,
    celdas: Tuple[str, str, str, str],
    fortalezas: Tuple[str, str, str, str]
) -> None:
    """
    Escribe cada fortaleza en una celda específica no consecutiva.

    Parámetros:
        hoja (Worksheet): Hoja de cálculo de openpyxl donde se escribirá.
        celdas (Tuple[str, str, str, str]): Coordenadas de celdas como strings (ej: "B7", "D9", ...).
        fortalezas (Tuple[str, str, str, str]): Tupla con los textos a escribir en las celdas.
    """

    for celda, texto in zip(celdas, fortalezas):
        anexar_texto(hoja, celda, texto, ancho_estimado=80)

#---------------------funcion para añadir celdas cuidado salud familiar-----------------------------

def escribir_duidado_por_hallazgos_en_celdas(
    hoja,
    celdas: Tuple[str, str, str, str],
    respuesta_hallazgos: Tuple[str, str, str, str]
) -> None:
    """
    Escribe cada fortaleza en una celda específica no consecutiva.

    Parámetros:
        hoja (Worksheet): Hoja de cálculo de openpyxl donde se escribirá.
        celdas (Tuple[str, str, str, str]): Coordenadas de celdas como strings (ej: "B7", "D9", ...).
        fortalezas (Tuple[str, str, str, str]): Tupla con los textos a escribir en las celdas.
    """
    anexar_texto(hoja, celdas[0], respuesta_hallazgos[0], ancho_estimado=50)
    anexar_texto(hoja, celdas[1], respuesta_hallazgos[1], ajustar_altura=False)
    anexar_texto(hoja, celdas[2], respuesta_hallazgos[2], ajustar_altura=False)
    anexar_texto(hoja, celdas[3], respuesta_hallazgos[3], ajustar_altura=False)


#------------------funcion para añadir celdas cuidado por curso de vida---------------
from typing import List, Tuple

def rellenar_celdas_integrantes(
    hoja: Worksheet,
    tuplas: List[Tuple[str, int, str, str, str, str]],
    ajustar_altura=True,
) -> None:
    """
    Rellena hasta 10 filas con información de tuplas. 
    Ajusta la altura de la fila SOLO según la longitud del campo de la columna 'G' (índice 2 de la tupla).

    :param hoja: Hoja de cálculo de openpyxl.
    :param tuplas: Lista de tuplas con 6 valores.
    :param ajustar_altura: Si se ajusta la altura (solo usando el campo G).
    :param ancho_estimado: Ancho estimado para calcular altura.
    """
    columnas = ["C", "F", "G", "L", "U", "V"]
    fila_inicial = 44
    max_filas = 10

    for i, datos in enumerate(tuplas[:max_filas]):
        fila = fila_inicial + i
        for idx, (col, valor) in enumerate(zip(columnas, datos)):
            ref = f"{col}{fila}"
            # Ajustamos altura solo si estamos en la columna G (índice 2)
            if idx == 2:
                anexar_texto(
                    hoja,
                    ref,
                    str(valor),
                    ajustar_altura=ajustar_altura,
                    ancho_estimado=30)
            else:
                anexar_texto(
                    hoja,
                    ref,
                    str(valor),
                    ajustar_altura=False  # No ajustar altura para otras columnas
                )

'''
from typing import List, Tuple


def rellenar_celdas_integrantes(
    hoja,
    tuplas: List[Tuple[str, int, str, str, str, str]],

) -> None:
    """
    Rellena hasta 10 filas en la hoja de Excel con la información de la lista de tuplas.
    Usa la función 'anexar_texto' para permitir control del texto y altura.

    :param hoja: Hoja de cálculo de openpyxl
    :param tuplas: Lista de tuplas con datos (nombre, edad, hallazgo, compromiso, logro trazador, logro intermedio)
    :param ajustar_altura: Si se desea ajustar la altura de las filas automáticamente.
    :param ancho_estimado: Ancho estimado para calcular la altura de la celda.
    """
    columnas = ["C", "F", "G", "L", "U", "V"]
    fila_inicial = 44
    max_filas = 10

    for i, datos in enumerate(tuplas[:max_filas]):
        fila = fila_inicial + i
        for col, valor in zip(columnas, datos):
            ref = f"{col}{fila}"
            anexar_texto(
                hoja,
                ref,
                str(valor),
                
            )
'''


#if __name__ == "__main__":



