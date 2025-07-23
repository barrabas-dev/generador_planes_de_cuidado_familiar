import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from situaciones_encontradas import llenar_situaciones_en_hoja, llenar_desarrollo_por_etapa, llenar_area_afectiva_en_hoja, llenar_area_personal_en_hoja, llenar_area_social_en_hoja, llenar_area_educativa_ocupacional_en_hoja, llenar_acciones_en_hoja, llenar_area_impresion_diagnostica_en_hoja, llenar_plan_intervencion_en_hoja, llenar_area_conclusion_en_hoja




def copiar_hoja(wb, hoja_origen, nuevo_nombre):
    """
    Crea una copia de una hoja dentro del libro con un nuevo nombre.
    """
    nueva_hoja = wb.copy_worksheet(hoja_origen)
    nueva_hoja.title = nuevo_nombre
    return nueva_hoja


def crear_planes_cuidado(dato_nucleos: dict, plantilla_path: str, salida_dir: str):
    """
    Genera archivos Excel por hogar con hojas individuales por persona seleccionada.

    :param dato_nucleos: Diccionario con hogares e integrantes.
    :param plantilla_path: Ruta al archivo plantilla Excel.
    :param salida_dir: Carpeta destino para los archivos generados.
    """

    if not os.path.isfile(plantilla_path):
        raise FileNotFoundError(f"Plantilla no encontrada: {plantilla_path}")

    os.makedirs(salida_dir, exist_ok=True)

    total_hojas_generadas = 0  # Contador de hojas

    for hogar_id, hogar_info in dato_nucleos.items():
        seleccionados = [
            p for p in hogar_info["integrantes"] if p.get("SELECCIONADO") is True
        ]

        if not seleccionados:
            continue

        # Crear archivo base desde plantilla
        hogar_filename = f"{hogar_id}.xlsx"
        hogar_path = os.path.join(salida_dir, hogar_filename)
        shutil.copy(plantilla_path, hogar_path)

        wb = load_workbook(hogar_path)
        hoja_modelo = wb.worksheets[0]  # Primera hoja como base

        # Procesar el primer seleccionado usando la hoja original
        primera_persona = seleccionados[0]
        hoja_modelo.title = primera_persona["NUMERO DE IDENTIFICACIÓN"]

        # Rellenar datos en la hoja base
        llenar_datos_en_hoja(hoja_modelo, primera_persona, hogar_info["integrantes"])
        insertar_imagenes_familiares(hoja_modelo, primera_persona["HOGAR"], "A21")
        insertar_texto_analisis(hoja_modelo, primera_persona["HOGAR"], "A23")
        total_hojas_generadas += 1  # Contar esta hoja


        # Para los demás seleccionados, duplicar la hoja base
        for persona in seleccionados[1:]:
            nueva_hoja = copiar_hoja(wb, hoja_modelo, persona["NUMERO DE IDENTIFICACIÓN"])
            llenar_datos_en_hoja(nueva_hoja, persona, hogar_info["integrantes"])
            insertar_imagenes_familiares(nueva_hoja, persona["HOGAR"], "A21")
            insertar_texto_analisis(nueva_hoja, persona["HOGAR"], "A23")
            total_hojas_generadas += 1  # Contar esta hoja


        wb.save(hogar_path)

    print(f"✅ Planes generados en: {salida_dir}")
    print(f"📄 Total de hojas generadas: {total_hojas_generadas}")


def llenar_datos_en_hoja(hoja, persona, todos_integrantes):
    """
    Llena los datos de una persona en la hoja correspondiente.
    Adaptar esta función para colocar los datos en las celdas deseadas.
    """
    hoy = datetime.today()
    hoja["B3"] = hoy.day
    hoja["C3"] = hoy.month
    hoja["D3"] = hoy.year


    hoja["G2"] = persona["HOGAR"]
    hoja["B5"] = persona["NOMBRE"]
    hoja["G5"] = persona["NUMERO DE IDENTIFICACIÓN"]
    hoja["B6"] = f'{persona["FECHA DE NACIMIENTO"].split()[0]} - Toledo'
    hoja["G6"] = persona["EDAD"]
    hoja["B7"] = persona["Direccion"]
    hoja["G7"] = persona["TELEFONO"]
    hoja["B25"] = "No Refiere"
    hoja["B26"] = "No Refiere"
    hoja["B28"] = "No Refiere"



    # --- Estructura Familiar ---
    fila_inicio = 10  # donde comienza la primera fila de datos, justo debajo de los encabezados

    for i, integrante in enumerate(todos_integrantes):
        fila = fila_inicio + i
        hoja[f"A{fila}"] = integrante["NOMBRE"]
        hoja[f"B{fila}"] = integrante["EDAD"]
        hoja[f"C{fila}"] = integrante.get("PARENTESCO", "")  # dejar en blanco si no existe
        hoja[f"D{fila}"] = integrante["ESTADO CIVIL"]
        hoja[f"E{fila}"] = integrante["ESCOLARIDAD"]
        hoja[f"F{fila}"] = integrante["OCUPACION"]
        hoja[f"G{fila}"] = integrante["CONVIVE DENTRO DE LA CASA"]

    #----celdas de situaciones encontradas---
    llenar_situaciones_en_hoja(hoja, persona, celda="B27")
    
    #----celda de desarrollo por etapas-------
    llenar_desarrollo_por_etapa(hoja, persona, celda="B29")

    #---celda de area personal-----------
    llenar_area_personal_en_hoja(hoja, persona, celda="B31")
    
    #---celda de area afectiva-----------
    llenar_area_afectiva_en_hoja(hoja, persona, celda="B34")
    
    #---celda de area social-----------
    llenar_area_social_en_hoja(hoja, persona, celda="B32")

    #---celda de area educativa/ocupacional-----------
    llenar_area_educativa_ocupacional_en_hoja(hoja, persona, celda="B33")

    #---celda de procesos de evaluacion-----------
    llenar_acciones_en_hoja(hoja, persona, celda="A36")

    #----celda de impresion diagnostica------------
    llenar_area_impresion_diagnostica_en_hoja(hoja, persona, celda="A38")

    #----celda de plan de intervencion-------------
    llenar_plan_intervencion_en_hoja(hoja, persona, celda="A40")

    #----celda de colclusiones y recomendaciones--------
    llenar_area_conclusion_en_hoja(hoja, persona, celda="A42")





#------------------------------------------------------------------------------------------------------------------------

from openpyxl.drawing.image import Image
from PIL import Image as PILImage

def insertar_imagenes_familiares(hoja, hogar_id, celda_inicio, ancho_celda_px=800):
    """
    Inserta dos imágenes (ecomapa y familiograma) en la hoja en la misma fila,
    una a la izquierda y otra a la derecha, partiendo desde la celda combinada.

    :param hoja: objeto worksheet
    :param hogar_id: identificador del hogar (ej. "H0005")
    :param celda_inicio: celda combinada donde deben ir las imágenes (ej. "B30")
    :param ancho_celda_px: ancho total estimado en píxeles de la celda combinada
    """

    # Rutas de las imágenes
    ruta_ecomapa = f"assets/imagenes/ecomapas/{hogar_id}.jpg"
    ruta_familiograma = f"assets/imagenes/familiogramas/{hogar_id}.jpg"

    # Posición de destino para las imágenes
    col, fila = celda_inicio[0], int(celda_inicio[1:])

    # Establecer altura de la fila
    hoja.row_dimensions[fila].height = 310


    if os.path.exists(ruta_ecomapa):
        img_eco = Image(ruta_ecomapa)
        img_eco.width = ancho_celda_px // 2  # mitad del ancho
        img_eco.height = int(img_eco.width * 1)  # proporción
        hoja.add_image(img_eco, f"{col}{fila}")
    
    if os.path.exists(ruta_familiograma):
        # Calcular una columna desplazada a la derecha
        col_derecha = chr(ord(col) + 3)  # Ajusta este valor según el tamaño de celda
        img_fam = Image(ruta_familiograma)
        img_fam.width = ancho_celda_px // 2
        img_fam.height = int(img_fam.width * 1)
        hoja.add_image(img_fam, f"{col_derecha}{fila}")

#------------------------------------------------------------------------------------------------------------------------------------
def insertar_texto_analisis(hoja, hogar_id, celda_inicio, ancho_max_caracteres=100):
    """
    Inserta el texto de análisis en una celda, ajustando el alto de la fila según la longitud del texto.

    :param hoja: objeto worksheet de openpyxl
    :param hogar_id: ID del hogar, usado para formar el nombre del archivo .txt (ej: "H0005")
    :param celda_inicio: celda donde se insertará el texto (ej: "B30")
    :param ancho_max_caracteres: número aproximado de caracteres por línea (para estimar altura)
    """
    ruta_texto = f"assets/analisis/{hogar_id}.txt"
    
    if not os.path.exists(ruta_texto):
        print(f"Texto de análisis no encontrado: {ruta_texto}")
        return

    # Leer contenido del archivo .txt
    with open(ruta_texto, "r", encoding="utf-8") as f:
        contenido = f.read().strip()

    # Insertar texto en la celda correspondiente
    hoja[celda_inicio] = contenido

    # Ajustar altura de fila basada en cantidad de líneas estimadas
    fila = int(''.join(filter(str.isdigit, celda_inicio)))
    num_lineas_estimadas = max(1, len(contenido) // ancho_max_caracteres + 1)
    hoja.row_dimensions[fila].height = num_lineas_estimadas * 15  # Ajuste básico

    # (Opcional) Establecer ajuste de texto en la celda
    hoja[celda_inicio].alignment = hoja[celda_inicio].alignment.copy(wrap_text=True)








'''
def insertar_imagen_analisis(hoja, hogar_id, celda_inicio, ancho_celda_px=800):
    """
    Inserta una imagen de análisis en una celda ajustando su tamaño al ancho de celda definido
    y modificando la altura de la fila para adaptarse al alto de la imagen.

    :param hoja: objeto worksheet de openpyxl
    :param hogar_id: ID del hogar, usado para formar el nombre del archivo (ej: "H0005")
    :param celda_inicio: celda donde se insertará la imagen (ej: "B30")
    :param ancho_celda_px: ancho estimado en píxeles de la celda combinada
    """
    ruta_imagen = f"assets/imagenes/analisis/{hogar_id}.jpg"
    
    if not os.path.exists(ruta_imagen):
        print(f"Imagen no encontrada: {ruta_imagen}")
        return

    # Cargar imagen original con PIL
    with PILImage.open(ruta_imagen) as img:
        ancho_original, alto_original = img.size
        escala = ancho_celda_px / ancho_original
        nuevo_alto = int(alto_original * escala)

    # Crear imagen para openpyxl
    img_excel = Image(ruta_imagen)
    img_excel.width = ancho_celda_px
    img_excel.height = nuevo_alto

    # Calcular fila y columna desde la celda
    col = celda_inicio[0]
    fila = int(celda_inicio[1:])

    # Ajustar altura de la fila en puntos (aproximadamente 0.75 puntos = 1 píxel)
    hoja.row_dimensions[fila].height = nuevo_alto * 0.75

    # Insertar imagen en la hoja
    hoja.add_image(img_excel, celda_inicio)

'''