# excel_exporter.py

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import shutil


class ExcelExporter:
    """
    Clase para generar hojas de cálculo a partir de una plantilla Excel.
    """

    def __init__(self, template_path: str, output_dir: str = "output"):
        """
        Inicializa el exportador con la ruta de la plantilla y el directorio de salida.

        :param template_path: Ruta a la plantilla .xlsx
        :param output_dir: Directorio donde se guardarán los archivos generados
        """
        self.template_path = Path(template_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        if not self.template_path.exists():
            raise FileNotFoundError(f"La plantilla '{self.template_path}' no existe.")

    def export(self, filename: str, data: dict, sheet_name: str = None) -> str:
        """
        Genera una nueva hoja de cálculo basada en la plantilla, rellenando los datos especificados.

        :param filename: Nombre del archivo de salida (sin extensión)
        :param data: Diccionario con los datos a escribir. Formato: {'A1': 'valor1', 'B2': 'valor2', ...}
        :param sheet_name: Nombre de la hoja donde se insertarán los datos. Si es None, se usa la activa.
        :return: Ruta al archivo generado
        """
        output_file = self.output_dir / f"{filename}.xlsx"

        # Copiar la plantilla a un nuevo archivo
        shutil.copy(self.template_path, output_file)

        # Cargar el archivo copiado
        wb = openpyxl.load_workbook(output_file)

        # Seleccionar la hoja adecuada
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"La hoja '{sheet_name}' no existe en la plantilla.")
            ws = wb[sheet_name]
        else:
            ws = wb.active  # Usa la hoja activa si no se especifica otra

        # Insertar los valores proporcionados
        for cell, value in data.items():
            if not isinstance(cell, str):
                raise TypeError(f"La clave '{cell}' debe ser una referencia de celda (por ejemplo, 'A1').")
            ws[cell] = value

        # Guardar el archivo modificado
        wb.save(output_file)

        return str(output_file)