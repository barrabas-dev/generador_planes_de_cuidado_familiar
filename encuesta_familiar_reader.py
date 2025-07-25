
from typing import Dict
from dataclasses import dataclass
from datetime import datetime
import openpyxl

# Lista de columnas booleanas
BOOLEAN_FIELDS = [
    "COMUNICACION NO ASERTIVA, USO INADECUADO DE PANTALLAS LO CUAL DIFICULTA LAS INTERACCIONES FAMILIARES",
    "INASISTENCIA A SERVICIOS DE SALUD DENTRO DEL MARCO DE LOS SERVICIOS DE LA RUTA DE PROMOCION Y MANTENIMIENTO DE LA SALUD",
    "RIESGO DE AISLAMIENTO SOCIAL Y DETERIORO EN SU SALUD FÍSICA Y MENTAL DEBIDO A LA FALTA DE APOYO FAMILIAR CONSTANTE Y HÁBITOS IRREGULARES DE AUTOCUIDADO",
    "ENVEJECIMIENTO, ENFERMEDADES CRÓNICAS Y CAMBIOS EMOCIONALES",
    "RIESGO DE MAL NUTRICIÓN",
    "ACTIVIDADES DEPORTIVAS COMUNITARIAS",
    "ACTIVIDADES LÚDICAS Y CULTURALES COMUNITARIAS",
    "RED DE APOYO COMUNITARIO",
    "SOBRECARGA DE CUIDADOR"
]


@dataclass
class EvaluacionFamiliar:
    fecha_visita: tuple
    resultado_apgar: str
    factores: Dict[str, bool]


def interpretar_bool(valor: str) -> bool:
    if valor is None:
        return False
    return str(valor).strip().lower() == "x"

'''
def formatear_fecha(valor) -> str:
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    return str(valor).strip()


'''
def formatear_y_dividir_fecha(valor) -> tuple[int, int, int]:
    """
    Recibe una fecha tipo datetime o string y devuelve (día, mes, año)
    """
    if not isinstance(valor, datetime):
        valor = datetime.strptime(str(valor).strip(), "%Y-%m-%d")
    return valor.day, valor.month, valor.year



def leer_encuesta_familiar(path_excel: str, hoja: str = None) -> Dict[str, EvaluacionFamiliar]:
    wb = openpyxl.load_workbook(path_excel, data_only=True)
    ws = wb[hoja] if hoja else wb.active

    encabezados = [cell.value for cell in ws[2]]
    resultados: Dict[str, EvaluacionFamiliar] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        # Validación anticipada: si no hay núcleo familiar, omitir
        idx_nucleo = encabezados.index("NUCLEO FAMILIAR") if "NUCLEO FAMILIAR" in encabezados else -1
        if idx_nucleo == -1 or not row[idx_nucleo] or str(row[idx_nucleo]).strip() == "":
            continue

        # Crear diccionario fila con normalización de contenido
        fila = {
            encabezado: str(valor).strip() if isinstance(valor, str) else valor
            for encabezado, valor in zip(encabezados, row)
        }

        nucleo_id = str(fila.get("NUCLEO FAMILIAR", "")).strip()
        fecha = formatear_y_dividir_fecha(fila.get("FECHA DE LA VISITA", ""))
        apgar = str(fila.get("RESULTADO DEL APGAR", "")).strip()

        factores = {
            campo: interpretar_bool(fila.get(campo)) for campo in BOOLEAN_FIELDS
        }

        evaluacion = EvaluacionFamiliar(
            fecha_visita=fecha,
            resultado_apgar=apgar,
            factores=factores
        )

        resultados[nucleo_id] = evaluacion

    return resultados



#----------------------------------------------------------------

#test-funtion
#'''
if __name__ == "__main__":
  # Ruta al archivo Excel que quieres probar
    ruta_excel = "assets/encuesta_familiar.xlsx"  # cámbialo por el nombre real si es distinto

    # Llamamos la función
    evaluaciones = leer_encuesta_familiar(ruta_excel)

    # Imprimimos los resultados
    for nucleo_id, evaluacion in evaluaciones.items():
        print(f"\n--- Núcleo Familiar: {nucleo_id} ---")
        print(f"Fecha de la Visita: {evaluacion.fecha_visita}")
        print(f"Resultado del APGAR: {evaluacion.resultado_apgar}")
        print("Factores identificados:")
        for factor, valor in evaluacion.factores.items():
            print(f"  {factor}: {valor}")
#'''


